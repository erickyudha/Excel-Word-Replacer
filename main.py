import os
import re
from openpyxl import load_workbook

# Dictionary for word replacement, reordered to prevent partial replacements
word_replacement_dict = {
    "Fasilkom-TI": "FN8",
    "FMIPA": "FN16",
    "F Hut": "FN6",
    "FFarm": "FN14",
    "FKG": "FN17",
    "FIB": "FN10",
    "FISIP": "FN1",
    "FKEP": "FN11",
    "FKM": "FN13",
    "FK": "FN4",
    "FPsi": "FN12",
    "FP": "FN5",
    "FEB": "FN3",
    "FV": "FN2",
    "FH": "FN7",
    "FT": "FN15",
    "SPS": "FN9",
}

# Regular expression pattern to match whole words
word_pattern = re.compile(r'\b(' + '|'.join(re.escape(word) for word in word_replacement_dict.keys()) + r')\b')

def replace_words_in_cell(cell_value):
    # Replace whole words based on the dictionary
    return word_pattern.sub(lambda x: word_replacement_dict[x.group()], cell_value)

def process_excel_files(folder_path):
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder_path, file_name)
            print("Processing file:", file_name)
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            cell.value = replace_words_in_cell(cell.value)
            wb.save(file_path)
            print("File", file_name, "processed successfully.")

# Provide the folder path containing Excel files
folder_path = "./output"
process_excel_files(folder_path)
